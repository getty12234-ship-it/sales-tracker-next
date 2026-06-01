# -*- coding: utf-8 -*-
"""セールストラッカーの日報動画スケジュール(st_monthly_videos)をxlsxに出力"""
import urllib.request, json, datetime, os

# st_monthly_videos の現在の全データ（date,videoId）※ダミー削除済み・7月分追加済み
DATA = """
2026-05-26,H7fjKcJ6CnA
2026-05-27,M4RO5I1Uor0
2026-05-28,M3YcN0YIkU4
2026-05-29,nKzNXbwuMMg
2026-06-01,6dEtGI8NSH8
2026-06-02,015NkD-Kb3E
2026-06-03,aOPLkFvo_Io
2026-06-04,j_UFGOPmLJc
2026-06-05,CdmCZtf4BMY
2026-06-08,-kdrwwFQWYw
2026-06-09,_sZvkdft44Y
2026-06-10,HLT0ECYsXrA
2026-06-11,P5E2L0FFzVs
2026-06-12,2hn4NgeO-qw
2026-06-15,ieUptV5chh4
2026-06-16,D_N9ZKMJBEI
2026-06-17,FoZ-490jTRI
2026-06-18,5uylkb4bwFM
2026-06-19,m_oVEGzlVbk
2026-06-22,cy9y4sAHRFY
2026-06-23,H7fjKcJ6CnA
2026-06-24,e_5dM0Ai-Sc
2026-06-25,auppeUC7RoQ
2026-06-26,pCfj25jIGD4
2026-06-29,W0bX_fv5Xg0
2026-06-30,jgXmhW98PPI
2026-07-01,aOPLkFvo_Io
2026-07-02,j_UFGOPmLJc
2026-07-03,CdmCZtf4BMY
2026-07-06,6dEtGI8NSH8
2026-07-07,015NkD-Kb3E
2026-07-08,HLT0ECYsXrA
2026-07-09,P5E2L0FFzVs
2026-07-10,2hn4NgeO-qw
2026-07-13,-kdrwwFQWYw
2026-07-14,_sZvkdft44Y
2026-07-15,FoZ-490jTRI
2026-07-16,5uylkb4bwFM
2026-07-17,m_oVEGzlVbk
2026-07-20,ieUptV5chh4
2026-07-21,D_N9ZKMJBEI
2026-07-22,e_5dM0Ai-Sc
2026-07-23,auppeUC7RoQ
2026-07-24,pCfj25jIGD4
2026-07-27,cy9y4sAHRFY
2026-07-28,jgXmhW98PPI
2026-07-29,M4RO5I1Uor0
2026-07-30,M3YcN0YIkU4
2026-07-31,nKzNXbwuMMg
""".strip()

WD = ['月', '火', '水', '木', '金', '土', '日']
THEME = {0: 'マインドセット', 1: 'アポ獲得', 2: 'テスクロ', 3: 'ラポール', 4: '成約・振り返り', 5: '—', 6: '—'}
THEME_COLOR = {
    'マインドセット': 'FDE9D9', 'アポ獲得': 'DCE6F1', 'テスクロ': 'E4DFEC',
    'ラポール': 'F2DCDB', '成約・振り返り': 'EBF1DE', '—': 'F2F2F2',
}
MONTH_FILL = {5: 'FFF7E6', 6: 'EAF3FB', 7: 'F0F7EA'}  # 月ごとの薄い背景


def get_title(vid):
    for _ in range(2):
        try:
            u = f"https://www.youtube.com/oembed?url=https://www.youtube.com/watch?v={vid}&format=json"
            req = urllib.request.Request(u, headers={'User-Agent': 'Mozilla/5.0'})
            with urllib.request.urlopen(req, timeout=10) as r:
                return json.loads(r.read().decode('utf-8')).get('title', '(タイトル不明)')
        except Exception:
            continue
    return '(タイトル取得不可)'


rows = []
for line in DATA.splitlines():
    ds, vid = line.split(',')
    d = datetime.date.fromisoformat(ds)
    print(f"取得中 {ds} ...", flush=True)
    rows.append((ds, WD[d.weekday()], THEME[d.weekday()], get_title(vid),
                 f"https://www.youtube.com/watch?v={vid}", d.month))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "日報動画スケジュール"

ws.merge_cells('A1:E1')
ws['A1'] = f'Sales Tracker 日報動画スケジュール（全{len(rows)}本）'
ws['A1'].font = Font(bold=True, size=14, color='FFFFFF')
ws['A1'].fill = PatternFill('solid', fgColor='4472C4')
ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
ws.row_dimensions[1].height = 26

headers = ['日付', '曜日', 'テーマ', '動画タイトル', '動画URL']
hfill = PatternFill('solid', fgColor='305496')
thin = Side(style='thin', color='BFBFBF')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
for c, h in enumerate(headers, 1):
    cell = ws.cell(row=2, column=c, value=h)
    cell.fill = hfill
    cell.font = Font(bold=True, color='FFFFFF')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border

for i, (ds, wd, theme, title, url, month) in enumerate(rows):
    r = i + 3
    for c, v in enumerate([ds, wd, theme, title, url], 1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.border = border
        cell.alignment = Alignment(vertical='center',
                                   horizontal='center' if c in (1, 2, 3) else 'left',
                                   wrap_text=(c == 4))
        if c in (1, 2):
            cell.fill = PatternFill('solid', fgColor=MONTH_FILL.get(month, 'FFFFFF'))
    ws.cell(row=r, column=3).fill = PatternFill('solid', fgColor=THEME_COLOR.get(theme, 'F2F2F2'))
    ucell = ws.cell(row=r, column=5)
    ucell.hyperlink = url
    ucell.font = Font(color='0563C1', underline='single')

for c, w in enumerate([12, 6, 14, 62, 46], 1):
    ws.column_dimensions[get_column_letter(c)].width = w
ws.freeze_panes = 'A3'

out_dir = r"C:\Users\81908\claude code\my-life\income\株式会社CoCo\05_チーム管理部門"
out = os.path.join(out_dir, "セールストラッカー日報動画スケジュール.xlsx")
try:
    wb.save(out)
    print("\nSAVED:", out)
except PermissionError:
    out = os.path.join(out_dir, "セールストラッカー日報動画スケジュール_最新.xlsx")
    wb.save(out)
    print("\nLOCKED -> SAVED_AS:", out)
print("件数:", len(rows))
